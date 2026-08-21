"""Excel export: one sheet normally, split sheets past the row ceiling."""

from openpyxl import load_workbook

from spse import export_excel


def write_csv(path, rows):
    lines = ["kode|nama"] + [f"{i}|paket {i}" for i in range(rows)]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    return path


def test_writes_single_sheet_with_header(tmp_path):
    csv_path = write_csv(tmp_path / "kemkes_2025_tender.csv", 5)

    xlsx = export_excel(csv_path, log=lambda *a: None)

    assert xlsx == csv_path.with_suffix(".xlsx")
    book = load_workbook(xlsx, read_only=True)
    assert book.sheetnames == ["data"]
    rows = list(book["data"].values)
    assert rows[0] == ("kode", "nama")
    assert len(rows) == 6


def test_splits_into_extra_sheets_past_the_row_limit(tmp_path):
    # Same behaviour as the real 1_000_000 cap, shrunk so the test is cheap.
    csv_path = write_csv(tmp_path / "kemkes_2025_tender.csv", 7)

    xlsx = export_excel(csv_path, log=lambda *a: None, rows_per_sheet=3)

    book = load_workbook(xlsx, read_only=True)
    assert book.sheetnames == ["data", "data_2", "data_3"]
    # Every sheet repeats the header; data rows are split 3/3/1.
    for name, expected in (("data", 3), ("data_2", 3), ("data_3", 1)):
        rows = list(book[name].values)
        assert rows[0] == ("kode", "nama")
        assert len(rows) == expected + 1
    assert list(book["data_3"].values)[1] == ("6", "paket 6")


def test_progress_reports_every_csv_row_once(tmp_path):
    csv_path = write_csv(tmp_path / "kemkes_2025_tender.csv", 4)
    seen = []

    export_excel(csv_path, log=lambda *a: None, rows_per_sheet=2,
                 progress=lambda done, total: seen.append((done, total)))

    assert seen[-1] == (5, 5)
